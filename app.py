import os
import shutil
import traceback
from datetime import datetime
from functools import wraps
from num2words import num2words

from flask import (Flask, flash, redirect, render_template,
                   request, send_file, session, jsonify)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from database import init_db, get_db

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "payroll-super-secret-2024-xK9mP")

EXPORT_DIR = "/tmp/payroll_exports"
BACKUP_DIR = "/tmp/payroll_backups"
os.makedirs(EXPORT_DIR, exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)

init_db()


@app.errorhandler(500)
def internal_error(error):
    tb = traceback.format_exc()
    return f"""<html><body style='font-family:monospace;padding:30px;background:#1e293b;color:#f8fafc'>
    <h2 style='color:#ef4444'>500 Error</h2><pre style='background:#0f172a;padding:20px;border-radius:8px;overflow:auto'>{tb}</pre>
    <a href='/' style='color:#60a5fa'>← Home</a></body></html>""", 500


# ── Helpers ──────────────────────────────────────────────────
def amount_in_words(amount):
    try:
        rupees = int(amount)
        paise  = round((amount - rupees) * 100)
        text   = num2words(rupees, lang='en_IN').title()
        if paise:
            text += f" And {num2words(paise, lang='en_IN').title()} Paise"
        return text + " Only"
    except Exception:
        return ""


def pt_slab(gross):
    if gross <= 15000: return 0
    if gross <= 20000: return 150
    return 200


# ── Decorators ───────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "user" not in session:
            return redirect("/login")
        return f(*args, **kwargs)
    return wrapper


def role_required(allowed_roles):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if "user" not in session:
                return redirect("/login")
            if session.get("role") not in allowed_roles:
                return render_template("403.html"), 403
            return f(*args, **kwargs)
        return wrapper
    return decorator


# ── Health ───────────────────────────────────────────────────
@app.route("/health")
def health():
    try:
        conn   = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM users")
        uc = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM employees")
        ec = cursor.fetchone()[0]
        conn.close()
        return {"status": "ok", "users": uc, "employees": ec}, 200
    except Exception as e:
        return {"status": "error", "detail": str(e)}, 500


# ── Auth ─────────────────────────────────────────────────────
@app.route("/login", methods=["GET", "POST"])
def login():
    if "user" in session:
        return redirect("/")
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        conn   = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users WHERE username=? AND password=?", (username, password))
        user = cursor.fetchone()
        conn.close()
        if user:
            session.permanent = True
            session["user"] = user["username"]
            session["role"] = user["role"]
            flash(f"Welcome back, {user['username']}!", "success")
            return redirect("/")
        error = "Invalid username or password."
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")


# ── Dashboard ────────────────────────────────────────────────
@app.route("/")
@login_required
def dashboard():
    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM employees WHERE status='Active'")
    emp_count = cursor.fetchone()[0]
    cursor.execute("SELECT COALESCE(SUM(net),0) FROM payroll")
    payout = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM payroll")
    processed = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM leaves WHERE status='Pending'")
    pending_leaves = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM overtime")
    ot_count = cursor.fetchone()[0]
    cursor.execute("""SELECT e.name, e.department, e.designation, p.net, p.month
        FROM payroll p JOIN employees e ON e.id=p.emp_id
        ORDER BY p.id DESC LIMIT 6""")
    recent_payroll = cursor.fetchall()
    cursor.execute("""SELECT e.name, l.from_date, l.to_date, l.leave_type, l.days
        FROM leaves l JOIN employees e ON e.id=l.emp_id
        ORDER BY l.id DESC LIMIT 5""")
    recent_leaves = cursor.fetchall()
    conn.close()
    today = datetime.now().strftime("%A, %d %B %Y")
    return render_template("dashboard.html",
                           emp_count=emp_count, payout=float(payout),
                           processed=processed, pending_leaves=pending_leaves,
                           ot_count=ot_count, recent_payroll=recent_payroll,
                           recent_leaves=recent_leaves, today=today)


# ── Employees ────────────────────────────────────────────────
@app.route("/employees")
@role_required(["admin", "hr"])
def employees():
    conn   = get_db()
    cursor = conn.cursor()
    q = request.args.get("q", "")
    dept = request.args.get("dept", "")
    sql = "SELECT * FROM employees WHERE 1=1"
    params = []
    if q:
        sql += " AND (name LIKE ? OR emp_code LIKE ?)"
        params += [f"%{q}%", f"%{q}%"]
    if dept:
        sql += " AND department=?"
        params.append(dept)
    sql += " ORDER BY id DESC"
    cursor.execute(sql, params)
    emps = cursor.fetchall()
    cursor.execute("SELECT id, name FROM companies ORDER BY name")
    companies = cursor.fetchall()
    cursor.execute("SELECT DISTINCT department FROM employees WHERE department!='' ORDER BY department")
    depts = [r[0] for r in cursor.fetchall()]
    conn.close()
    return render_template("employees.html", employees=emps, companies=companies, depts=depts, q=q, dept=dept)


@app.route("/add_employee", methods=["POST"])
@role_required(["admin", "hr"])
def add_employee():
    f = request.form
    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO employees
            (company_id, emp_code, name, department, designation, location, doj, dob, gender,
             basic, hra, taxable_allow, night_shift_allow, pf_relief,
             bank_account, ifsc, pan, uan, pf_number, esi_number, eps_number, status)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        f.get("company_id", 1), f.get("emp_code", ""), f.get("name", ""),
        f.get("department", ""), f.get("designation", ""), f.get("location", ""),
        f.get("doj", ""), f.get("dob", ""), f.get("gender", "Male"),
        float(f.get("basic", 0)), float(f.get("hra", 0)),
        float(f.get("taxable_allow", 0)), float(f.get("night_shift_allow", 0)),
        float(f.get("pf_relief", 0)),
        f.get("bank_account", ""), f.get("ifsc", ""), f.get("pan", ""),
        f.get("uan", ""), f.get("pf_number", ""), f.get("esi_number", ""),
        f.get("eps_number", ""), "Active"
    ))
    conn.commit()
    conn.close()
    flash("Employee added successfully.", "success")
    return redirect("/employees")


@app.route("/edit_employee/<int:emp_id>", methods=["GET", "POST"])
@role_required(["admin", "hr"])
def edit_employee(emp_id):
    conn   = get_db()
    cursor = conn.cursor()
    if request.method == "POST":
        f = request.form
        cursor.execute("""
            UPDATE employees SET company_id=?, emp_code=?, name=?, department=?,
            designation=?, location=?, doj=?, dob=?, gender=?,
            basic=?, hra=?, taxable_allow=?, night_shift_allow=?, pf_relief=?,
            bank_account=?, ifsc=?, pan=?, uan=?, pf_number=?, esi_number=?,
            eps_number=?, status=? WHERE id=?
        """, (
            f.get("company_id", 1), f.get("emp_code", ""), f.get("name", ""),
            f.get("department", ""), f.get("designation", ""), f.get("location", ""),
            f.get("doj", ""), f.get("dob", ""), f.get("gender", "Male"),
            float(f.get("basic", 0)), float(f.get("hra", 0)),
            float(f.get("taxable_allow", 0)), float(f.get("night_shift_allow", 0)),
            float(f.get("pf_relief", 0)),
            f.get("bank_account", ""), f.get("ifsc", ""), f.get("pan", ""),
            f.get("uan", ""), f.get("pf_number", ""), f.get("esi_number", ""),
            f.get("eps_number", ""), f.get("status", "Active"), emp_id
        ))
        conn.commit()
        conn.close()
        flash("Employee updated.", "success")
        return redirect("/employees")
    cursor.execute("SELECT * FROM employees WHERE id=?", (emp_id,))
    emp = cursor.fetchone()
    cursor.execute("SELECT id, name FROM companies")
    companies = cursor.fetchall()
    conn.close()
    return render_template("edit_employee.html", emp=emp, companies=companies)


@app.route("/delete_employee/<int:emp_id>", methods=["POST"])
@role_required(["admin"])
def delete_employee(emp_id):
    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE employees SET status='Inactive' WHERE id=?", (emp_id,))
    conn.commit()
    conn.close()
    flash("Employee deactivated.", "warning")
    return redirect("/employees")


# ── Salary Structure ─────────────────────────────────────────
@app.route("/salary_structure")
@role_required(["admin", "hr"])
def salary_structure():
    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute("""SELECT e.id, e.emp_code, e.name, e.department, e.designation,
        e.basic, e.hra, e.taxable_allow, e.night_shift_allow, e.pf_relief
        FROM employees e WHERE e.status='Active' ORDER BY e.name""")
    emps = cursor.fetchall()
    conn.close()
    return render_template("salary_structure.html", employees=emps)


@app.route("/update_salary_structure/<int:emp_id>", methods=["POST"])
@role_required(["admin", "hr"])
def update_salary_structure(emp_id):
    f = request.form
    basic = float(f.get("basic", 0))
    hra   = float(f.get("hra", 0))
    ta    = float(f.get("taxable_allow", 0))
    nsa   = float(f.get("night_shift_allow", 0))
    pfr   = float(f.get("pf_relief", 0))
    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute("""UPDATE employees SET basic=?, hra=?, taxable_allow=?,
        night_shift_allow=?, pf_relief=? WHERE id=?""", (basic, hra, ta, nsa, pfr, emp_id))
    # Save history
    cursor.execute("""INSERT INTO salary_history
        (emp_id, effective_date, basic, hra, taxable_allow, night_shift_allow, pf_relief, remarks)
        VALUES (?,?,?,?,?,?,?,?)""",
        (emp_id, datetime.now().strftime("%Y-%m-%d"), basic, hra, ta, nsa, pfr,
         f.get("remarks", "Salary revision")))
    conn.commit()
    conn.close()
    flash("Salary structure updated.", "success")
    return redirect("/salary_structure")


# ── Attendance ───────────────────────────────────────────────
@app.route("/attendance", methods=["GET", "POST"])
@role_required(["admin", "hr"])
def attendance():
    conn   = get_db()
    cursor = conn.cursor()
    month = request.args.get("month", datetime.now().strftime("%Y-%m"))

    if request.method == "POST":
        emp_id       = request.form["emp_id"]
        m            = request.form["month"]
        days_in_month = int(request.form.get("days_in_month", 26))
        arrear_days  = int(request.form.get("arrear_days", 0))
        lopr_days    = int(request.form.get("lopr_days", 0))
        lop_days     = int(request.form.get("lop_days", 0))
        present_days = days_in_month + arrear_days + lopr_days - lop_days
        try:
            cursor.execute("""INSERT INTO attendance
                (emp_id, month, days_in_month, arrear_days, lopr_days, lop_days, present_days)
                VALUES (?,?,?,?,?,?,?)
                ON CONFLICT(emp_id, month) DO UPDATE SET
                days_in_month=excluded.days_in_month,
                arrear_days=excluded.arrear_days,
                lopr_days=excluded.lopr_days,
                lop_days=excluded.lop_days,
                present_days=excluded.present_days""",
                (emp_id, m, days_in_month, arrear_days, lopr_days, lop_days, present_days))
            conn.commit()
            flash("Attendance saved.", "success")
        except Exception as e:
            flash(f"Error: {e}", "danger")
        conn.close()
        return redirect(f"/attendance?month={m}")

    cursor.execute("SELECT id, name FROM employees WHERE status='Active' ORDER BY name")
    emps = cursor.fetchall()
    cursor.execute("""SELECT a.*, e.name FROM attendance a
        JOIN employees e ON e.id=a.emp_id WHERE a.month=? ORDER BY e.name""", (month,))
    records = cursor.fetchall()
    conn.close()
    return render_template("attendance.html", employees=emps, records=records, month=month)


@app.route("/delete_attendance/<int:att_id>", methods=["POST"])
@role_required(["admin", "hr"])
def delete_attendance(att_id):
    conn = get_db()
    conn.execute("DELETE FROM attendance WHERE id=?", (att_id,))
    conn.commit()
    conn.close()
    flash("Attendance record deleted.", "warning")
    return redirect("/attendance")


# ── Leaves ───────────────────────────────────────────────────
@app.route("/leave", methods=["GET", "POST"])
@role_required(["admin", "hr"])
def leave():
    conn   = get_db()
    cursor = conn.cursor()
    if request.method == "POST":
        f = request.form
        from_date = f.get("from_date", "")
        to_date   = f.get("to_date", from_date)
        try:
            d1 = datetime.strptime(from_date, "%Y-%m-%d")
            d2 = datetime.strptime(to_date, "%Y-%m-%d")
            days = max(1, (d2 - d1).days + 1)
        except Exception:
            days = 1
        cursor.execute("""INSERT INTO leaves (emp_id, from_date, to_date, leave_type, days, reason, status)
            VALUES (?,?,?,?,?,?,?)""",
            (f["emp_id"], from_date, to_date, f.get("leave_type", "CL"),
             days, f.get("reason", ""), f.get("status", "Approved")))
        conn.commit()
        conn.close()
        flash("Leave recorded.", "success")
        return redirect("/leave")

    cursor.execute("SELECT id, name FROM employees WHERE status='Active' ORDER BY name")
    emps = cursor.fetchall()
    cursor.execute("""SELECT l.*, e.name FROM leaves l
        JOIN employees e ON e.id=l.emp_id ORDER BY l.id DESC""")
    records = cursor.fetchall()
    conn.close()
    return render_template("leave.html", employees=emps, records=records)


@app.route("/edit_leave/<int:leave_id>", methods=["POST"])
@role_required(["admin", "hr"])
def edit_leave(leave_id):
    f = request.form
    conn = get_db()
    conn.execute("UPDATE leaves SET leave_type=?, status=?, reason=? WHERE id=?",
        (f.get("leave_type"), f.get("status"), f.get("reason"), leave_id))
    conn.commit()
    conn.close()
    flash("Leave updated.", "success")
    return redirect("/leave")


@app.route("/delete_leave/<int:leave_id>", methods=["POST"])
@role_required(["admin", "hr"])
def delete_leave(leave_id):
    conn = get_db()
    conn.execute("DELETE FROM leaves WHERE id=?", (leave_id,))
    conn.commit()
    conn.close()
    flash("Leave deleted.", "warning")
    return redirect("/leave")


# ── Overtime ─────────────────────────────────────────────────
@app.route("/overtime", methods=["GET", "POST"])
@role_required(["admin", "hr"])
def overtime():
    conn   = get_db()
    cursor = conn.cursor()
    if request.method == "POST":
        f    = request.form
        hours = float(f.get("hours", 0))
        rate  = float(f.get("rate", 0))
        amount = hours * rate
        ot_date = f.get("ot_date", "")
        month = ot_date[:7] if len(ot_date) >= 7 else ""
        cursor.execute("""INSERT INTO overtime (emp_id, ot_date, hours, rate, amount, month)
            VALUES (?,?,?,?,?,?)""",
            (f["emp_id"], ot_date, hours, rate, amount, month))
        conn.commit()
        conn.close()
        flash("Overtime recorded.", "success")
        return redirect("/overtime")

    cursor.execute("SELECT id, name FROM employees WHERE status='Active' ORDER BY name")
    emps = cursor.fetchall()
    cursor.execute("""SELECT o.*, e.name FROM overtime o
        JOIN employees e ON e.id=o.emp_id ORDER BY o.ot_date DESC""")
    records = cursor.fetchall()
    conn.close()
    return render_template("overtime.html", employees=emps, records=records)


@app.route("/edit_overtime/<int:ot_id>", methods=["POST"])
@role_required(["admin", "hr"])
def edit_overtime(ot_id):
    f = request.form
    hours = float(f.get("hours", 0))
    rate  = float(f.get("rate", 0))
    conn = get_db()
    conn.execute("UPDATE overtime SET hours=?, rate=?, amount=? WHERE id=?",
                 (hours, rate, hours * rate, ot_id))
    conn.commit()
    conn.close()
    flash("Overtime updated.", "success")
    return redirect("/overtime")


@app.route("/delete_overtime/<int:ot_id>", methods=["POST"])
@role_required(["admin", "hr"])
def delete_overtime(ot_id):
    conn = get_db()
    conn.execute("DELETE FROM overtime WHERE id=?", (ot_id,))
    conn.commit()
    conn.close()
    flash("Overtime deleted.", "warning")
    return redirect("/overtime")


# ── Payroll ──────────────────────────────────────────────────
@app.route("/payroll", methods=["GET", "POST"])
@role_required(["admin", "hr", "accountant"])
def payroll():
    conn   = get_db()
    cursor = conn.cursor()
    month = request.args.get("month", datetime.now().strftime("%Y-%m"))

    if request.method == "POST":
        emp_id = int(request.form["emp_id"])
        month  = request.form["month"]

        cursor.execute("SELECT * FROM employees WHERE id=?", (emp_id,))
        emp = dict(cursor.fetchone())

        cursor.execute("""SELECT * FROM attendance WHERE emp_id=? AND month=?""", (emp_id, month))
        att = cursor.fetchone()
        days_in_month  = att["days_in_month"] if att else 26
        arrear_days    = att["arrear_days"]   if att else 0
        lopr_days      = att["lopr_days"]     if att else 0
        lop_days       = att["lop_days"]      if att else 0
        net_days       = days_in_month + arrear_days + lopr_days - lop_days

        # OT for this month
        cursor.execute("SELECT COALESCE(SUM(amount),0) FROM overtime WHERE emp_id=? AND month=?", (emp_id, month))
        ot_amount = float(cursor.fetchone()[0])

        # Proportional salary
        per_day  = (emp["basic"] + emp["hra"] + emp.get("taxable_allow", 0)
                    + emp.get("night_shift_allow", 0)) / days_in_month if days_in_month else 0
        lop_deduct = per_day * lop_days

        basic    = emp["basic"]
        hra      = emp["hra"]
        ta       = emp.get("taxable_allow", 0)
        nsa      = emp.get("night_shift_allow", 0)
        pfr      = emp.get("pf_relief", 0)
        gross    = basic + hra + ta + nsa + ot_amount + pfr - lop_deduct

        pf  = round(min(basic, 15000) * 0.12, 2)
        esi = round(gross * 0.0075, 2) if gross <= 21000 else 0
        pt  = pt_slab(gross)
        tds = float(request.form.get("tds", 0))
        other_ded = float(request.form.get("other_deduction", 0))
        net = gross - pf - esi - pt - tds - other_ded

        try:
            cursor.execute("""INSERT INTO payroll
                (emp_id, month, basic, hra, taxable_allow, night_shift_allow, ot_allowance,
                 pf_relief, gross, pf, esi, pt, tds, lop, other_deduction, net,
                 days_in_month, arrear_days, lopr_days, lop_days, net_days_worked)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(emp_id, month) DO UPDATE SET
                basic=excluded.basic, hra=excluded.hra, taxable_allow=excluded.taxable_allow,
                night_shift_allow=excluded.night_shift_allow, ot_allowance=excluded.ot_allowance,
                pf_relief=excluded.pf_relief, gross=excluded.gross,
                pf=excluded.pf, esi=excluded.esi, pt=excluded.pt,
                tds=excluded.tds, lop=excluded.lop, other_deduction=excluded.other_deduction,
                net=excluded.net, days_in_month=excluded.days_in_month,
                arrear_days=excluded.arrear_days, lopr_days=excluded.lopr_days,
                lop_days=excluded.lop_days, net_days_worked=excluded.net_days_worked""",
                (emp_id, month, basic, hra, ta, nsa, ot_amount, pfr,
                 gross, pf, esi, pt, tds, lop_deduct, other_ded, net,
                 days_in_month, arrear_days, lopr_days, lop_days, net_days))
            conn.commit()
            flash(f"Payroll processed for {emp['name']} — {month}.", "success")
        except Exception as e:
            flash(f"Error: {e}", "danger")
        conn.close()
        return redirect(f"/payroll?month={month}")

    cursor.execute("SELECT id, name FROM employees WHERE status='Active' ORDER BY name")
    emps = cursor.fetchall()
    cursor.execute("""SELECT p.*, e.name FROM payroll p
        JOIN employees e ON e.id=p.emp_id WHERE p.month=? ORDER BY e.name""", (month,))
    records = cursor.fetchall()
    conn.close()
    return render_template("payroll.html", employees=emps, records=records, month=month)


@app.route("/delete_payroll/<int:pay_id>", methods=["POST"])
@role_required(["admin"])
def delete_payroll(pay_id):
    conn = get_db()
    conn.execute("DELETE FROM payroll WHERE id=?", (pay_id,))
    conn.commit()
    conn.close()
    flash("Payroll record deleted.", "warning")
    return redirect("/payroll")


# ── Payslip PDF (matches sample format) ──────────────────────
@app.route("/payslip/<int:pay_id>")
@login_required
def payslip(pay_id):
    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT p.*, e.name, e.designation, e.department, e.emp_code,
               e.doj, e.dob, e.gender, e.location, e.pan, e.uan,
               e.bank_account, e.ifsc, e.pf_number, e.esi_number, e.eps_number,
               c.name as company_name, c.address, c.city, c.state,
               c.pf_number as co_pf, c.esi_number as co_esi
        FROM payroll p
        JOIN employees e ON e.id = p.emp_id
        JOIN companies c ON c.id = e.company_id
        WHERE p.id=?
    """, (pay_id,))
    row = cursor.fetchone()

    # YTD totals
    if row:
        emp_id = row["emp_id"]
        yr     = row["month"][:4]
        cursor.execute("""SELECT
            COALESCE(SUM(basic),0), COALESCE(SUM(hra),0),
            COALESCE(SUM(taxable_allow),0), COALESCE(SUM(night_shift_allow),0),
            COALESCE(SUM(ot_allowance),0), COALESCE(SUM(pf_relief),0),
            COALESCE(SUM(gross),0), COALESCE(SUM(pf),0), COALESCE(SUM(esi),0),
            COALESCE(SUM(pt),0), COALESCE(SUM(tds),0), COALESCE(SUM(lop),0), COALESCE(SUM(net),0)
            FROM payroll WHERE emp_id=? AND month LIKE ?""", (emp_id, f"{yr}%"))
        ytd = cursor.fetchone()
    conn.close()

    if not row:
        flash("Payslip not found.", "danger")
        return redirect("/payroll")

    d   = dict(row)
    ytd = list(ytd) if ytd else [0]*13
    month_label = d["month"]
    try:
        month_label = datetime.strptime(d["month"], "%Y-%m").strftime("%B %Y")
    except Exception:
        pass

    file_path = os.path.join(EXPORT_DIR, f"payslip_{pay_id}.pdf")

    # ── Build PDF with reportlab platypus ──
    doc = SimpleDocTemplate(file_path, pagesize=A4,
                            topMargin=10*mm, bottomMargin=10*mm,
                            leftMargin=12*mm, rightMargin=12*mm)
    styles = getSampleStyleSheet()
    story  = []
    W = A4[0] - 24*mm   # usable width

    def style(name, **kw):
        s = ParagraphStyle(name, parent=styles["Normal"], **kw)
        return s

    DARK  = colors.HexColor("#0f2557")
    LIGHT = colors.HexColor("#e8edf8")
    MID   = colors.HexColor("#d0d9f0")
    GREEN = colors.HexColor("#1a7a4a")

    # ── Header: Company + Title ──
    addr_str = " | ".join(filter(None, [
        d.get("address", ""), d.get("city", ""), d.get("state", "")]))
    header_data = [[
        Paragraph(f"""<font size="14"><b>{d.get('company_name','')}</b></font><br/>
            <font size="8" color="#555555">{addr_str}</font>""",
            style("hdr", leading=16)),
        Paragraph(f"""<b>SALARY PAYSLIP</b><br/>
            <font size="9">Month: {month_label}</font>""",
            style("title", fontSize=13, leading=18, alignment=TA_RIGHT))
    ]]
    header_tbl = Table(header_data, colWidths=[W*0.6, W*0.4])
    header_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), DARK),
        ("TEXTCOLOR",  (0,0), (-1,-1), colors.white),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING",(0,0), (-1,-1), 10),
        ("RIGHTPADDING",(0,0), (-1,-1), 10),
        ("TOPPADDING", (0,0), (-1,-1), 8),
        ("BOTTOMPADDING",(0,0),(-1,-1), 8),
    ]))
    story.append(header_tbl)
    story.append(Spacer(1, 3*mm))

    # ── Employee Info Grid ──
    co_pf  = d.get("co_pf") or d.get("pf_number", "")
    co_esi = d.get("co_esi") or d.get("esi_number", "")
    info_rows = [
        ["Emp Code", d.get("emp_code",""), "EMP Name", d.get("name",""), "PF No.", co_pf],
        ["Department", d.get("department",""), "Designation", d.get("designation",""), "ESI No.", co_esi],
        ["Location", d.get("location",""), "Bank A/c No", d.get("bank_account",""), "Pan No.", d.get("pan","")],
        ["Date of Birth", d.get("dob",""), "Gender", d.get("gender",""), "EPS No.", d.get("eps_number","")],
        ["Date of Joining", d.get("doj",""), "UAN", d.get("uan",""), "", ""],
    ]
    cw = [W*0.14, W*0.18, W*0.14, W*0.22, W*0.14, W*0.18]
    info_tbl = Table(info_rows, colWidths=cw)
    info_style = TableStyle([
        ("FONTSIZE",   (0,0), (-1,-1), 8),
        ("FONTNAME",   (0,0), (-1,-1), "Helvetica"),
        ("FONTNAME",   (0,0), (0,-1), "Helvetica-Bold"),
        ("FONTNAME",   (2,0), (2,-1), "Helvetica-Bold"),
        ("FONTNAME",   (4,0), (4,-1), "Helvetica-Bold"),
        ("BACKGROUND", (0,0), (0,-1), LIGHT),
        ("BACKGROUND", (2,0), (2,-1), LIGHT),
        ("BACKGROUND", (4,0), (4,-1), LIGHT),
        ("GRID",       (0,0), (-1,-1), 0.3, colors.HexColor("#aaaaaa")),
        ("TOPPADDING", (0,0), (-1,-1), 3),
        ("BOTTOMPADDING",(0,0),(-1,-1), 3),
        ("LEFTPADDING",(0,0), (-1,-1), 5),
    ])
    info_tbl.setStyle(info_style)
    story.append(info_tbl)
    story.append(Spacer(1, 3*mm))

    # ── Earnings & Deductions ──
    def fmt(v): return f"₹ {v:,.2f}" if v else "—"

    earn_hdr = [Paragraph("<b>Earnings</b>", style("eh", fontSize=8)),
                Paragraph("<b>Amount</b>", style("eh", fontSize=8, alignment=TA_RIGHT)),
                Paragraph("<b>YTD</b>",    style("eh", fontSize=8, alignment=TA_RIGHT)),
                Paragraph("<b>Deductions</b>", style("eh", fontSize=8)),
                Paragraph("<b>Amount</b>", style("eh", fontSize=8, alignment=TA_RIGHT)),
                Paragraph("<b>YTD</b>",    style("eh", fontSize=8, alignment=TA_RIGHT))]

    def rp(t, alignment=TA_LEFT): return Paragraph(t, style("cell", fontSize=8, alignment=alignment))
    def rv(v): return rp(fmt(v) if v else "—", TA_RIGHT)

    earn_rows = [
        [rp("Basic"),            rv(d["basic"]),           rv(ytd[0]),  rp("Provident Fund"),  rv(d["pf"]),  rv(ytd[7])],
        [rp("House Rent Allow."),rv(d["hra"]),             rv(ytd[1]),  rp("Professional Tax"),rv(d["pt"]),  rv(ytd[9])],
        [rp("Taxable Allowance"),rv(d["taxable_allow"]),   rv(ytd[2]),  rp("ESI"),             rv(d["esi"]), rv(ytd[8])],
        [rp("Night Shift Allow."),rv(d["night_shift_allow"]),rv(ytd[3]),rp("TDS"),             rv(d["tds"]), rv(ytd[10])],
        [rp("Over Time Allow."), rv(d["ot_allowance"]),    rv(ytd[4]),  rp("LOP Deduction"),   rv(d["lop"]), rv(ytd[11])],
        [rp("PF Relief"),        rv(d.get("pf_relief",0)), rv(ytd[5]),  rp("Other Deduction"), rv(d.get("other_deduction",0)), rp("—", TA_RIGHT)],
    ]
    total_earn = d["gross"]
    total_ded  = d["pf"] + d["esi"] + d["pt"] + d["tds"] + d["lop"] + d.get("other_deduction",0)
    total_row  = [
        Paragraph("<b>Total Earnings</b>", style("tot", fontSize=8, fontName="Helvetica-Bold")),
        Paragraph(f"<b>{fmt(total_earn)}</b>", style("tot", fontSize=8, alignment=TA_RIGHT, fontName="Helvetica-Bold")),
        rp(""),
        Paragraph("<b>Total Deductions</b>", style("tot", fontSize=8, fontName="Helvetica-Bold")),
        Paragraph(f"<b>{fmt(total_ded)}</b>",  style("tot", fontSize=8, alignment=TA_RIGHT, fontName="Helvetica-Bold")),
        rp(""),
    ]

    ed_cw = [W*0.22, W*0.1, W*0.1, W*0.22, W*0.1, W*0.1]  # Adjusted
    # Normalize to W total
    ed_cw = [W*0.225, W*0.11, W*0.11, W*0.225, W*0.105, W*0.105]

    ed_tbl = Table([earn_hdr] + earn_rows + [total_row], colWidths=ed_cw)
    ed_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), DARK),
        ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
        ("FONTSIZE",   (0,0), (-1,-1), 8),
        ("GRID",       (0,0), (-1,-1), 0.3, colors.HexColor("#aaaaaa")),
        ("BACKGROUND", (0,-1),(-1,-1), LIGHT),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0), (-1,-1), 3),
        ("BOTTOMPADDING",(0,0),(-1,-1), 3),
        ("LEFTPADDING",(0,0), (-1,-1), 5),
        ("LINEAFTER",  (2,0), (2,-1), 1, DARK),  # divider between earn/ded
    ]))
    story.append(ed_tbl)
    story.append(Spacer(1, 2*mm))

    # ── Net Pay ──
    net_data = [[
        Paragraph(f"<b>Net Pay Rs. {fmt(d['net'])}</b>",
                  style("np", fontSize=11, textColor=colors.white)),
        Paragraph(f"<b>In Words: {amount_in_words(d['net'])}</b>",
                  style("nw", fontSize=9, textColor=colors.white))
    ]]
    net_tbl = Table(net_data, colWidths=[W*0.3, W*0.7])
    net_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), GREEN),
        ("TOPPADDING", (0,0), (-1,-1), 6),
        ("BOTTOMPADDING",(0,0),(-1,-1), 6),
        ("LEFTPADDING",(0,0), (-1,-1), 10),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
    ]))
    story.append(net_tbl)
    story.append(Spacer(1, 2*mm))

    # ── Days Summary ──
    net_days = d.get("net_days_worked", 0) or (d["days_in_month"] + d["arrear_days"] + d["lopr_days"] - d["lop_days"])
    days_data = [
        [Paragraph("<b>Days in Month (A)</b>", style("dh", fontSize=8)),
         Paragraph("<b>Arrear Days (B)</b>", style("dh", fontSize=8)),
         Paragraph("<b>LOPR Days (C)</b>", style("dh", fontSize=8)),
         Paragraph("<b>LOP Days (D)</b>", style("dh", fontSize=8)),
         Paragraph("<b>Net Days Worked (E=A+B+C-D)</b>", style("dh", fontSize=8))],
        [rp(str(d.get("days_in_month","26"))), rp(str(d.get("arrear_days","0"))),
         rp(str(d.get("lopr_days","0"))), rp(str(d.get("lop_days","0"))),
         rp(str(net_days))],
    ]
    days_tbl = Table(days_data, colWidths=[W/5]*5)
    days_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), MID),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,-1), 8),
        ("GRID",       (0,0), (-1,-1), 0.3, colors.HexColor("#aaaaaa")),
        ("TOPPADDING", (0,0), (-1,-1), 3),
        ("BOTTOMPADDING",(0,0),(-1,-1), 3),
        ("LEFTPADDING",(0,0), (-1,-1), 5),
    ]))
    story.append(days_tbl)
    story.append(Spacer(1, 3*mm))

    # ── Footer ──
    footer_tbl = Table([[
        Paragraph("<i>This document contains confidential information. "
                  "Computer generated payslip — no signature required.</i>",
                  style("ft", fontSize=7, textColor=colors.grey, alignment=TA_CENTER))
    ]], colWidths=[W])
    footer_tbl.setStyle(TableStyle([
        ("TOPPADDING",(0,0),(-1,-1), 4),
        ("LINEABOVE", (0,0),(-1,-1), 0.5, colors.grey),
    ]))
    story.append(footer_tbl)

    doc.build(story)

    return send_file(file_path, as_attachment=True,
                     download_name=f"Payslip_{d['name']}_{d['month']}.pdf")


# ── Reports ──────────────────────────────────────────────────
@app.route("/reports")
@role_required(["admin", "hr", "accountant"])
def reports():
    conn   = get_db()
    cursor = conn.cursor()
    month = request.args.get("month", datetime.now().strftime("%Y-%m"))
    cursor.execute("""SELECT e.name, e.department, e.designation,
        p.basic, p.hra, p.taxable_allow, p.night_shift_allow, p.ot_allowance,
        p.gross, p.pf, p.esi, p.pt, p.tds, p.lop, p.net
        FROM payroll p JOIN employees e ON e.id=p.emp_id
        WHERE p.month=? ORDER BY e.name""", (month,))
    records = cursor.fetchall()
    totals  = {
        "gross": sum(r["gross"] for r in records),
        "pf":    sum(r["pf"]    for r in records),
        "esi":   sum(r["esi"]   for r in records),
        "pt":    sum(r["pt"]    for r in records),
        "tds":   sum(r["tds"]   for r in records),
        "net":   sum(r["net"]   for r in records),
    }
    conn.close()
    return render_template("reports.html", records=records, month=month, totals=totals)


@app.route("/salary_register")
@role_required(["admin", "hr", "accountant"])
def salary_register():
    conn   = get_db()
    cursor = conn.cursor()
    month = request.args.get("month", datetime.now().strftime("%Y-%m"))
    cursor.execute("""SELECT e.emp_code, e.name, e.department, e.designation,
        e.bank_account, p.*
        FROM payroll p JOIN employees e ON e.id=p.emp_id
        WHERE p.month=? ORDER BY e.name""", (month,))
    records = cursor.fetchall()
    conn.close()
    return render_template("salary_register.html", records=records, month=month)


@app.route("/salary_register_excel")
@role_required(["admin", "hr", "accountant"])
def salary_register_excel():
    conn   = get_db()
    cursor = conn.cursor()
    month = request.args.get("month", datetime.now().strftime("%Y-%m"))
    cursor.execute("""SELECT e.emp_code, e.name, e.department, e.designation,
        e.bank_account, p.basic, p.hra, p.taxable_allow, p.night_shift_allow,
        p.ot_allowance, p.pf_relief, p.gross, p.pf, p.esi, p.pt, p.tds, p.lop, p.net
        FROM payroll p JOIN employees e ON e.id=p.emp_id
        WHERE p.month=? ORDER BY e.name""", (month,))
    rows = cursor.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Salary Register"
    hdr_fill = PatternFill("solid", fgColor="0F2557")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    headers  = ["Emp Code","Name","Department","Designation","Bank A/c",
                "Basic","HRA","Tax Allow","Night Shift","OT Allow","PF Relief",
                "Gross","PF","ESI","PT","TDS","LOP","Net Pay"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(list(row))
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 14

    path = os.path.join(EXPORT_DIR, f"salary_register_{month}.xlsx")
    wb.save(path)
    return send_file(path, as_attachment=True,
                     download_name=f"Salary_Register_{month}.xlsx")


@app.route("/bank_sheet")
@role_required(["admin", "hr", "accountant"])
def bank_sheet():
    conn   = get_db()
    cursor = conn.cursor()
    month = request.args.get("month", datetime.now().strftime("%Y-%m"))
    cursor.execute("""SELECT e.name, e.bank_account, e.ifsc, p.net
        FROM payroll p JOIN employees e ON e.id=p.emp_id
        WHERE p.month=? ORDER BY e.name""", (month,))
    data = cursor.fetchall()
    conn.close()
    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Sheet"
    hdr_fill = PatternFill("solid", fgColor="0F2557")
    hdr_font = Font(bold=True, color="FFFFFF")
    headers = ["S.No","Employee Name","Bank Account","IFSC Code","Net Amount"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
    for i, row in enumerate(data, 1):
        ws.append([i, row["name"], row["bank_account"], row["ifsc"], row["net"]])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    path = os.path.join(EXPORT_DIR, f"bank_sheet_{month}.xlsx")
    wb.save(path)
    return send_file(path, as_attachment=True,
                     download_name=f"Bank_Sheet_{month}.xlsx")


@app.route("/statutory_report")
@role_required(["admin", "accountant"])
def statutory_report():
    conn   = get_db()
    cursor = conn.cursor()
    month = request.args.get("month", datetime.now().strftime("%Y-%m"))
    cursor.execute("""SELECT e.name, e.uan, e.pf_number, e.esi_number,
        p.basic, p.pf, p.esi, p.pt
        FROM payroll p JOIN employees e ON e.id=p.emp_id
        WHERE p.month=? ORDER BY e.name""", (month,))
    records = cursor.fetchall()
    conn.close()
    return render_template("statutory_report.html", records=records, month=month)


# ── Salary History / Revision ─────────────────────────────────
@app.route("/salary_revision", methods=["GET", "POST"])
@role_required(["admin", "hr"])
def salary_revision():
    conn   = get_db()
    cursor = conn.cursor()
    if request.method == "POST":
        f = request.form
        cursor.execute("""INSERT INTO salary_history
            (emp_id, effective_date, basic, hra, taxable_allow, night_shift_allow, pf_relief, remarks)
            VALUES (?,?,?,?,?,?,?,?)""",
            (f["emp_id"], f["date"],
             float(f.get("basic",0)), float(f.get("hra",0)),
             float(f.get("taxable_allow",0)), float(f.get("night_shift_allow",0)),
             float(f.get("pf_relief",0)), f.get("remarks","")))
        # Update employee current salary
        cursor.execute("""UPDATE employees SET basic=?, hra=?, taxable_allow=?,
            night_shift_allow=?, pf_relief=? WHERE id=?""",
            (float(f.get("basic",0)), float(f.get("hra",0)),
             float(f.get("taxable_allow",0)), float(f.get("night_shift_allow",0)),
             float(f.get("pf_relief",0)), f["emp_id"]))
        conn.commit()
        conn.close()
        flash("Salary revision saved and applied.", "success")
        return redirect("/salary_revision")
    cursor.execute("SELECT id, name FROM employees WHERE status='Active' ORDER BY name")
    emps = cursor.fetchall()
    cursor.execute("""SELECT sh.*, e.name FROM salary_history sh
        JOIN employees e ON e.id=sh.emp_id ORDER BY sh.effective_date DESC LIMIT 30""")
    history = cursor.fetchall()
    conn.close()
    return render_template("salary_revision.html", employees=emps, history=history)


@app.route("/salary_history/<int:emp_id>")
@role_required(["admin", "hr", "accountant"])
def salary_history(emp_id):
    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM employees WHERE id=?", (emp_id,))
    emp  = cursor.fetchone()
    cursor.execute("""SELECT * FROM salary_history WHERE emp_id=?
        ORDER BY effective_date DESC""", (emp_id,))
    records = cursor.fetchall()
    conn.close()
    return render_template("salary_history.html", records=records, emp=emp)


# ── Companies ─────────────────────────────────────────────────
@app.route("/companies", methods=["GET"])
@role_required(["admin"])
def companies():
    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM companies ORDER BY id")
    data = cursor.fetchall()
    conn.close()
    return render_template("companies.html", companies=data)


@app.route("/add_company", methods=["POST"])
@role_required(["admin"])
def add_company():
    f = request.form
    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute("""INSERT INTO companies
        (name, address, city, state, pincode, phone, email, pf_number, esi_number, pan, tan)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
        (f["name"], f.get("address",""), f.get("city",""), f.get("state",""),
         f.get("pincode",""), f.get("phone",""), f.get("email",""),
         f.get("pf",""), f.get("esi",""), f.get("pan",""), f.get("tan","")))
    conn.commit()
    conn.close()
    flash("Company added.", "success")
    return redirect("/companies")


@app.route("/edit_company/<int:cid>", methods=["POST"])
@role_required(["admin"])
def edit_company(cid):
    f = request.form
    conn = get_db()
    conn.execute("""UPDATE companies SET name=?, address=?, city=?, state=?,
        pincode=?, phone=?, email=?, pf_number=?, esi_number=?, pan=?, tan=? WHERE id=?""",
        (f["name"], f.get("address",""), f.get("city",""), f.get("state",""),
         f.get("pincode",""), f.get("phone",""), f.get("email",""),
         f.get("pf",""), f.get("esi",""), f.get("pan",""), f.get("tan",""), cid))
    conn.commit()
    conn.close()
    flash("Company updated.", "success")
    return redirect("/companies")


@app.route("/delete_company/<int:cid>", methods=["POST"])
@role_required(["admin"])
def delete_company(cid):
    conn = get_db()
    conn.execute("DELETE FROM companies WHERE id=?", (cid,))
    conn.commit()
    conn.close()
    flash("Company deleted.", "warning")
    return redirect("/companies")


# ── Users ─────────────────────────────────────────────────────
@app.route("/users")
@role_required(["admin"])
def users():
    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM users ORDER BY id")
    data = cursor.fetchall()
    conn.close()
    return render_template("users.html", users=data)


@app.route("/add_user", methods=["POST"])
@role_required(["admin"])
def add_user():
    f = request.form
    conn   = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO users (username, password, role, email) VALUES (?,?,?,?)",
                       (f["username"], f["password"], f["role"], f.get("email","")))
        conn.commit()
        flash("User added.", "success")
    except Exception as e:
        flash(f"Error: {e}", "danger")
    conn.close()
    return redirect("/users")


@app.route("/edit_user/<int:uid>", methods=["POST"])
@role_required(["admin"])
def edit_user(uid):
    f = request.form
    conn = get_db()
    if f.get("password"):
        conn.execute("UPDATE users SET username=?, password=?, role=?, email=? WHERE id=?",
                     (f["username"], f["password"], f["role"], f.get("email",""), uid))
    else:
        conn.execute("UPDATE users SET username=?, role=?, email=? WHERE id=?",
                     (f["username"], f["role"], f.get("email",""), uid))
    conn.commit()
    conn.close()
    flash("User updated.", "success")
    return redirect("/users")


@app.route("/delete_user/<int:uid>", methods=["POST"])
@role_required(["admin"])
def delete_user(uid):
    if uid == 1:
        flash("Cannot delete primary admin.", "danger")
        return redirect("/users")
    conn = get_db()
    conn.execute("DELETE FROM users WHERE id=?", (uid,))
    conn.commit()
    conn.close()
    flash("User deleted.", "warning")
    return redirect("/users")


# ── Backup ────────────────────────────────────────────────────
@app.route("/backup")
@role_required(["admin"])
def backup():
    from database import DB_PATH
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_file = os.path.join(BACKUP_DIR, f"payroll_{timestamp}.db")
    shutil.copy(DB_PATH, backup_file)
    flash(f"Backup created: payroll_{timestamp}.db", "success")
    return redirect("/backup_page")


@app.route("/restore/<filename>")
@role_required(["admin"])
def restore(filename):
    from database import DB_PATH
    safe_name   = os.path.basename(filename)
    backup_file = os.path.join(BACKUP_DIR, safe_name)
    if not os.path.exists(backup_file):
        flash("Backup file not found.", "danger")
        return redirect("/backup_page")
    shutil.copy(backup_file, DB_PATH)
    flash("Database restored successfully.", "success")
    return redirect("/backup_page")


@app.route("/backup_page")
@role_required(["admin"])
def backup_page():
    files = sorted(os.listdir(BACKUP_DIR), reverse=True) if os.path.exists(BACKUP_DIR) else []
    return render_template("backup.html", files=files)


if __name__ == "__main__":
    app.run(debug=True)
